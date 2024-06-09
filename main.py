# Imports #
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.remote.webelement import WebElement
import time
from os import path
from openpyxl import load_workbook


# Drop images logic
JS_DROP_FILES = """
var k=arguments,d=k[0],g=k[1],c=k[2],m=d.ownerDocument||document;for(var e=0;;){var f=d.getBoundingClientRect(),b=f.left+(g||(f.width/2)),a=f.top+(c||(f.height/2)),h=m.elementFromPoint(b,a);
if(h&&d.contains(h)){break}if(++e>1){var j=new Error('Element not interactable');j.code=15;throw j}d.scrollIntoView({behavior:'instant',block:'center',inline:'center'})}var l=m.createElement('INPUT');
l.setAttribute('type','file');l.setAttribute('multiple','');l.setAttribute('style','position:fixed;z-index:2147483647;left:0;top:0;');l.onchange=function(q){l.parentElement.removeChild(l);q.stopPropagation();
var r={constructor:DataTransfer,effectAllowed:'all',dropEffect:'none',types:['Files'],files:l.files,setData:function u(){},getData:function o(){},clearData:function s(){},setDragImage:function i(){}};
if(window.DataTransferItemList){r.items=Object.setPrototypeOf(Array.prototype.map.call(l.files,function(x){return{constructor:DataTransferItem,kind:'file',type:x.type,getAsFile:function v(){return x},getAsString:function y(A){var z=new FileReader();z.onload=function(B){A(B.target.result)};
z.readAsText(x)},webkitGetAsEntry:function w(){return{constructor:FileSystemFileEntry,name:x.name,fullPath:'/'+x.name,isFile:true,isDirectory:false,file:function z(A){A(x)}}}}}),{constructor:DataTransferItemList,add:function t(){},clear:function p(){},remove:function n(){}})}['dragenter','dragover','drop'].forEach(function(v){var w=m.createEvent('DragEvent');
w.initMouseEvent(v,true,true,m.defaultView,0,0,0,b,a,false,false,false,false,0,null);Object.setPrototypeOf(w,null);w.dataTransfer=r;Object.setPrototypeOf(w,DragEvent.prototype);h.dispatchEvent(w)})};m.documentElement.appendChild(l);l.getBoundingClientRect();return l
"""


def drop_files(element, files, offsetX=0, offsetY=0):
    driver = element.parent
    isLocal = not driver._is_remote or "127.0.0.1" in driver.command_executor._url
    paths = []
    for file in files if isinstance(files, list) else [files]:
        if not path.isfile(file):
            raise FileNotFoundError(file)
        paths.append(file if isLocal else element._upload(file))

    value = "\n".join(paths)
    elm_input = driver.execute_script(JS_DROP_FILES, element, offsetX, offsetY)
    elm_input._execute("sendKeysToElement", {"value": [value], "text": value})


# Important to link it to DOM
WebElement.drop_files = drop_files


# Read data from excel
def read_excel_tabs(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    list_of_dicts = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        image, title, price, category, condition, color, description, location = row
        if location:
            entry_dict = {
                "Image": image,
                "Title": title,
                "Price": price,
                "Category": category,
                "Condition": condition,
                "Color": color,
                "Description": description,
                "Location": location,
            }
            list_of_dicts.append(entry_dict)
    return list_of_dicts


# Read data from excel
def read_profiles(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    list_of_dicts = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        email, password = row
        if email:
            entry_dict = {"Email": email, "Password": password}
            list_of_dicts.append(entry_dict)
    return list_of_dicts


# Check if all images exist
def images_exist(data):
    i = 0
    no_image_missing = True
    for tab in data:
        i += 1
        if not path.exists(tab["Image"]):
            no_image_missing = False
            print("Tab# " + str(i) + " : Image not found : " + tab["Image"])
    return no_image_missing


# Open Chrome
def open_browser():
    options = Options()
    options.add_argument("--disable-notifications")
    driver = webdriver.Chrome(
        service=Service("drivers/chromedriver.exe"), options=options
    )
    return driver


# Read settings from settings file
def read_settings():
    data = open("config.txt", "r").read().split("\n")
    config = {}
    for line in data:
        if line == "" or line.startswith("#"):
            continue
        line = line.replace("'", "")
        key, value = line.split(" = ")
        config[key] = value
    return config


# Log into facebook using creds
def login_facebook(driver, actions, profile):
    driver.find_element(By.ID, "email").click()
    actions.send_keys(profile["Email"]).perform()
    time.sleep(0.5)
    driver.find_element(By.ID, "pass").click()
    actions.send_keys(profile["Password"]).perform()
    time.sleep(1)
    driver.find_element(
        By.XPATH, "//button[contains(@data-testid,'royal_login_button')]"
    ).click()


# Open Tabs of new listing
def open_tabs(driver, count):
    url = "https://www.facebook.com/marketplace/create/item/"
    driver.get(url)
    for i in range(1, count):
        driver.execute_script("window.open('{}', '_blank');".format(url))


# Check if condition is same
def is_same_condition(text, tab_text):
    text = text.split()
    tab_text = tab_text.split()
    if text[0] == tab_text[0] and text[-1] == tab_text[-1]:
        return True
    return False


# Submit as fast as possible
def submit_quickly(driver):
    i = 0
    for handle in driver.window_handles:
        i += 1
        driver.switch_to.window(handle)
        # Publish
        try:
            publish_button = driver.find_element(
                By.XPATH, "//div[contains(@aria-label,'Publish')]"
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", publish_button)
            time.sleep(0.2)
            publish_button.click()
        except:
            try:
                # Focus as well
                publish_button = driver.find_element(
                    By.XPATH, "//div[contains(@aria-label,'Publish')]"
                )
                driver.execute_script(
                    "arguments[0].scrollIntoView(true);", publish_button
                )
                driver.execute_script(
                    "arguments[0].setAttribute('aria-hidden', 'false')", publish_button
                )
                driver.execute_script(
                    "arguments[0].setAttribute('tabindex', '0')", publish_button
                )
                driver.execute_script("arguments[0].focus();", publish_button)
                driver.execute_script("arguments[0].click();", publish_button)
                actions.move_to_element(next_button).click(next_button).perform()
                time.sleep(0.2)
                publish_button.click()
            except:
                pass
        print("Published Tab#", i)


# Fill content
def fill_data(driver, actions, tab):
    # Image
    buttons = driver.find_elements(By.XPATH, "//div[contains(@role,'button')]")
    image_drop = None
    more_details_element = None
    for button in buttons:
        try:
            if (
                "photo" in button.text.lower()
                and "drag and drop" in button.text.lower()
            ):
                image_drop = button
            if "more details" in button.text.lower():
                more_details_element = button
        except:
            pass

    image_drop.drop_files(tab["Image"])
    # Title
    title_element = driver.find_element(
        By.XPATH, "//label[contains(@aria-label,'Title')]"
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", title_element)
    time.sleep(0.2)
    title_element.click()
    actions.send_keys(tab["Title"]).perform()
    # Price
    driver.find_element(By.XPATH, "//label[contains(@aria-label,'Price')]").click()
    actions.send_keys(tab["Price"]).perform()
    # Category
    cat_elem = driver.find_element(
        By.XPATH, "//label[contains(@aria-label,'Category')]"
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", cat_elem)
    time.sleep(0.2)
    cat_elem.click()
    time.sleep(3)
    elements = driver.find_elements(By.XPATH, "//div[contains(@role,'button')]")
    for elem in elements:
        try:
            if str(elem.text.strip()).lower() == tab["Category"].strip().lower():
                elem.click()
                break
        except:
            pass
    # Condition
    con_elem = driver.find_element(
        By.XPATH, "//label[contains(@aria-label,'Condition')]"
    )
    con_elem.click()
    time.sleep(3)
    elements = driver.find_elements(By.XPATH, "//div[contains(@role,'option')]")
    for elem in elements:
        try:
            if is_same_condition(
                str(elem.text.strip()).lower(), tab["Condition"].strip().lower()
            ):
                elem.click()
                break
        except:
            pass
    # Color
    if tab["Color"] != "" and tab["Color"] is not None:
        color_element = driver.find_element(
            By.XPATH, "//label[contains(@aria-label,'Color')]"
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", color_element)
        color_element.click()
        actions.send_keys(tab["Color"]).perform()
    # Description
    if tab["Description"] != "" and tab["Description"] is not None:
        desc_element = driver.find_element(
            By.XPATH, "//label[contains(@aria-label,'Description')]"
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", desc_element)
        desc_element.click()
        actions.send_keys(tab["Description"]).perform()
    # Availability
    avail_elem = driver.find_element(
        By.XPATH, "//label[contains(@aria-label,'Availability')]"
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", avail_elem)
    time.sleep(0.2)
    avail_elem.click()
    time.sleep(1)
    elements = driver.find_elements(By.XPATH, "//div[contains(@role,'option')]")
    for elem in elements:
        try:
            if "list as in stock" in str(elem.text.strip()).lower():
                elem.click()
                break
        except:
            pass
    # Location
    try:
        loc_element = driver.find_element(
            By.XPATH, "//label[contains(@aria-label,'Location')]"
        )
    except:
        driver.execute_script(
            "arguments[0].scrollIntoView(true);", more_details_element
        )
        more_details_element.click()
        time.sleep(0.5)
        loc_element = driver.find_element(
            By.XPATH, "//label[contains(@aria-label,'Location')]"
        )
    driver.execute_script("arguments[0].scrollIntoView(true);", loc_element)
    time.sleep(0.2)
    loc_element.click()
    actions.send_keys(tab["Location"]).perform()
    time.sleep(2)
    try:
        location_list_parent = driver.find_element(
            By.XPATH, "//ul[contains(@role,'listbox')]"
        )
        location_list = location_list_parent.find_elements(
            By.XPATH, "//li[contains(@role,'option')]"
        )
        location_list[0].click()
    except:
        time.sleep(3)
        location_list_parent = driver.find_element(
            By.XPATH, "//ul[contains(@role,'listbox')]"
        )
        location_list = location_list_parent.find_elements(
            By.XPATH, "//li[contains(@role,'option')]"
        )
        location_list[0].click()
    # Checkboxes
    try:
        cells = driver.find_elements(By.XPATH, "//div[contains(@role,'checkbox')]")
        for cell in cells:
            cell.click()
    except:
        pass
    time.sleep(1)
    # Next
    try:
        next_button = driver.find_element(
            By.XPATH, "//div[contains(@aria-label,'Next')]"
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
        next_button.click()
    except:
        try:
            publish_button = driver.find_element(
                By.XPATH, "//div[contains(@aria-label,'Publish')]"
            )
        except:
            next_button = driver.find_element(
                By.XPATH, "//div[contains(@aria-label,'Next')]"
            )
            try:
                # Focus as well
                driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
                driver.execute_script(
                    "arguments[0].setAttribute('aria-hidden', 'false')", next_button
                )
                driver.execute_script(
                    "arguments[0].setAttribute('tabindex', '0')", next_button
                )
                driver.execute_script("arguments[0].focus();", next_button)
                driver.execute_script("arguments[0].click();", next_button)
                actions.move_to_element(next_button).click(next_button).perform()
                time.sleep(0.2)
                next_button.click()
            except:
                pass
    time.sleep(0.2)


# Main Control Flow
if __name__ == "__main__":
    config = read_settings()
    data = read_excel_tabs(config["tabs_file"])
    profiles = read_profiles(config["profiles_file"])
    if len(profiles) == 0:
        print("\n\nInsert Profiles in profiles.xlsx file to Run.\n")
        time.sleep(20)
        quit()
    if not images_exist(data):
        time.sleep(20)
        quit()
    drivers = []
    for profile in profiles:
        print(
            "\nUsing Profile ", profile["Email"], " ->  Running for", len(data), "tabs."
        )
        driver = open_browser()
        actions = ActionChains(driver)
        driver.get("https://www.facebook.com/")
        time.sleep(5)
        login_facebook(driver, actions, profile)
        time.sleep(3)
        open_tabs(driver, len(data))
        time.sleep(3)
        # Fill data
        for i in range(0, len(data)):
            print("Filling Data for Tab#", str(i + 1))
            driver.switch_to.window(driver.window_handles[i])
            time.sleep(0.5)
            fill_data(driver, actions, data[i])
        submit_quickly(driver)
        time.sleep(25)
        driver.quit()
